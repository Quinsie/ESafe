import asyncio
from datetime import date
from io import BytesIO
from typing import Any, Literal

from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy import text
from sqlalchemy.ext.asyncio import AsyncEngine

RegionLevel = Literal["SIDO", "SIGUNGU"]
TopPercent = Literal[1, 5, 10]

EXPORT_HEADERS = (
    "번호",
    "건물명",
    "지번주소",
    "광역시도",
    "시군구",
    "지역 내 위험순위",
    "광주·전남 전체 위험순위",
    "위험점수",
    "상위백분위(%)",
    "건물 주용도",
    "최근 점검·검사일",
    "6개월 내 점검·검사 여부",
    "1년 내 점검·검사 여부",
    "점검·검사 이력 건수",
)

LEVEL_NAMES: dict[RegionLevel, str] = {
    "SIDO": "광역시·도",
    "SIGUNGU": "시·군·구",
}

def _hierarchy_cte() -> str:
    return """
        WITH building_hierarchy AS (
            SELECT b.building_id,
                   local.region_code AS local_code,
                   local.level AS local_level,
                   CASE
                     WHEN local.level = 'SIGUNGU' THEN local.region_code
                     WHEN local.level = 'EUPMYEONDONG' THEN local.parent_code
                   END AS sigungu_code,
                   CASE
                     WHEN local.level = 'SIDO' THEN local.region_code
                     ELSE sido.region_code
                   END AS sido_code,
                   sigungu.name AS sigungu_name,
                   CASE
                     WHEN local.level = 'SIDO' THEN local.name
                     ELSE sido.name
                   END AS sido_name
            FROM building b
            JOIN admin_region local ON local.region_code = b.region_code
            LEFT JOIN admin_region sigungu
              ON sigungu.region_code = CASE
                   WHEN local.level = 'SIGUNGU' THEN local.region_code
                   WHEN local.level = 'EUPMYEONDONG' THEN local.parent_code
                 END
            LEFT JOIN admin_region sido
              ON sido.region_code = CASE
                   WHEN local.level = 'SIDO' THEN local.region_code
                   ELSE sigungu.parent_code
                 END
        )
    """


async def extract_region_options(
    engine: AsyncEngine,
    level: RegionLevel,
    parent_code: str | None,
    timeout_seconds: float,
) -> dict[str, Any]:
    async def query() -> dict[str, Any]:
        region_code = {
            "SIDO": "hierarchy.sido_code",
            "SIGUNGU": "hierarchy.sigungu_code",
        }[level]
        parent_filter = (
            "AND region.parent_code = :parent_code" if parent_code is not None else ""
        )
        statement = text(
            _hierarchy_cte()
            + f"""
            SELECT region.region_code, region.name, region.full_name, region.parent_code,
                   count(*) AS building_count,
                   count(*) FILTER (WHERE risk.top_percentile <= 1) AS top_1_count,
                   count(*) FILTER (WHERE risk.top_percentile <= 5) AS top_5_count,
                   count(*) FILTER (WHERE risk.top_percentile <= 10) AS top_10_count
            FROM building_hierarchy hierarchy
            JOIN admin_region region ON region.region_code = {region_code}
            JOIN building_risk_snapshot risk
              ON risk.building_id = hierarchy.building_id
             AND risk.reference_month = DATE '2026-03-01'
             AND risk.horizon_days = 60
             AND risk.lineage_version = 'v27.1-focus-2026-03-60d'
            WHERE region.level = :level
              {parent_filter}
              AND EXISTS (
                SELECT 1 FROM reference_dataset_state WHERE state_id = true
              )
            GROUP BY region.region_code, region.name, region.full_name, region.parent_code
            ORDER BY region.full_name, region.region_code
            """
        )
        async with engine.connect() as connection:
            parameters: dict[str, str] = {"level": level}
            if parent_code is not None:
                parameters["parent_code"] = parent_code
            rows = (
                await connection.execute(
                    statement,
                    parameters,
                )
            ).mappings().all()
        return {
            "level": level,
            "levelName": LEVEL_NAMES[level],
            "items": [
                {
                    "regionCode": row["region_code"],
                    "name": row["name"],
                    "fullName": row["full_name"],
                    "parentCode": row["parent_code"],
                    "buildingCount": int(row["building_count"]),
                    "eligibleCounts": {
                        "1": int(row["top_1_count"]),
                        "5": int(row["top_5_count"]),
                        "10": int(row["top_10_count"]),
                    },
                }
                for row in rows
            ],
            "riskReference": {
                "referenceMonth": "2026-03",
                "horizonDays": 60,
                "lineageVersion": "v27.1-focus-2026-03-60d",
                "isProbability": False,
            },
        }

    return await asyncio.wait_for(query(), timeout=max(timeout_seconds, 10))


async def extract_building_rows(
    engine: AsyncEngine,
    level: RegionLevel,
    region_code: str,
    top_percent: TopPercent,
    timeout_seconds: float,
) -> tuple[str, list[dict[str, Any]]]:
    async def query() -> tuple[str, list[dict[str, Any]]]:
        filter_column = {
            "SIDO": "hierarchy.sido_code",
            "SIGUNGU": "hierarchy.sigungu_code",
        }[level]
        statement = text(
            _hierarchy_cte()
            + f"""
            , selected AS (
                SELECT b.building_id, b.building_name, b.lot_address,
                       b.customer_data ->> 'main_use_name' AS main_use_name,
                       hierarchy.sido_name, hierarchy.sigungu_name,
                       risk.final_score, risk.regional_rank,
                       risk.top_percentile,
                       row_number() OVER (
                         ORDER BY risk.final_score DESC, b.building_id
                       ) AS selected_region_rank
                FROM building_hierarchy hierarchy
                JOIN building b ON b.building_id = hierarchy.building_id
                JOIN building_risk_snapshot risk
                  ON risk.building_id = b.building_id
                 AND risk.reference_month = DATE '2026-03-01'
                 AND risk.horizon_days = 60
                 AND risk.lineage_version = 'v27.1-focus-2026-03-60d'
                WHERE {filter_column} = :region_code
                  AND EXISTS (
                    SELECT 1 FROM reference_dataset_state WHERE state_id = true
                  )
            )
            SELECT selected.*,
                   facility.latest_inspection_date,
                   coalesce(facility.inspection_record_count, 0) AS inspection_record_count
            FROM selected
            LEFT JOIN LATERAL (
                SELECT max(entity.last_inspection_date) AS latest_inspection_date,
                       coalesce(sum(entity.source_row_count), 0) AS inspection_record_count
                FROM building_facility_link link
                JOIN facility_entity entity ON entity.facility_id = link.facility_id
                WHERE link.building_id = selected.building_id
            ) facility ON true
            WHERE selected.top_percentile <= :top_percent
            ORDER BY selected.selected_region_rank, selected.building_id
            """
        )
        async with engine.connect() as connection:
            region_name = (
                await connection.execute(
                    text(
                        """
                        SELECT full_name
                        FROM admin_region
                        WHERE region_code = :region_code AND level = :level
                        """
                    ),
                    {"region_code": region_code, "level": level},
                )
            ).scalar_one_or_none()
            if region_name is None:
                raise ValueError("region not found")
            rows = (
                await connection.execute(
                    statement,
                    {"region_code": region_code, "top_percent": top_percent},
                )
            ).mappings().all()
        return str(region_name), [dict(row) for row in rows]

    return await asyncio.wait_for(query(), timeout=max(timeout_seconds, 60))


def _months_ago(as_of: date, months: int) -> date:
    total_months = as_of.year * 12 + as_of.month - 1 - months
    year, month_index = divmod(total_months, 12)
    month = month_index + 1
    month_days = (31, 29 if year % 4 == 0 and (year % 100 != 0 or year % 400 == 0) else 28,
                  31, 30, 31, 30, 31, 31, 30, 31, 30, 31)
    return date(year, month, min(as_of.day, month_days[month - 1]))


def _inspection_flag(last_inspection_date: date | None, cutoff: date) -> str:
    if last_inspection_date is None:
        return "미등록"
    return "있음" if last_inspection_date >= cutoff else "없음"


def build_extract_workbook(
    *,
    level: RegionLevel,
    region_name: str,
    top_percent: TopPercent,
    rows: list[dict[str, Any]],
    as_of: date,
) -> bytes:
    workbook = Workbook(write_only=True)
    condition_sheet = workbook.create_sheet("추출 조건")
    condition_sheet.column_dimensions["A"].width = 22
    condition_sheet.column_dimensions["B"].width = 58
    condition_sheet.append(["항목", "내용"])
    for key, value in (
        ("추출 기준일", as_of.isoformat()),
        ("지역 단위", LEVEL_NAMES[level]),
        ("선택 지역", region_name),
        ("위험도 범위", f"광주·전남 모델 상위 {top_percent}%"),
        ("기준 위험도", "v27.1 · 2026-03 · 향후 60일 상대점수"),
        ("추출 건수", len(rows)),
        ("점검·검사 이력 기준", "연결 시설 원천의 최근 점검일 및 원천 행 수"),
    ):
        condition_sheet.append([key, value])

    sheet = workbook.create_sheet("건축물 목록")
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:N{max(2, len(rows) + 1)}"
    widths = (8, 24, 42, 16, 18, 16, 22, 14, 16, 14, 18, 24, 22, 20)
    for index, width in enumerate(widths, 1):
        sheet.column_dimensions[chr(64 + index)].width = width
    header_fill = PatternFill("solid", fgColor="173F70")
    header_font = Font(color="FFFFFF", bold=True)
    header_cells = []
    for value in EXPORT_HEADERS:
        cell = WriteOnlyCell(sheet, value=value)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        header_cells.append(cell)
    sheet.append(header_cells)

    six_month_cutoff = _months_ago(as_of, 6)
    one_year_cutoff = _months_ago(as_of, 12)
    for index, row in enumerate(rows, 1):
        last_inspection = row.get("latest_inspection_date")
        sheet.append(
            [
                index,
                row.get("building_name") or "건물명 미등록",
                row.get("lot_address") or "",
                row.get("sido_name") or "",
                row.get("sigungu_name") or "",
                int(row["selected_region_rank"]),
                int(row["regional_rank"]),
                float(row["final_score"]),
                float(row["top_percentile"]),
                row.get("main_use_name") or "미등록",
                last_inspection.isoformat() if last_inspection is not None else "",
                _inspection_flag(last_inspection, six_month_cutoff),
                _inspection_flag(last_inspection, one_year_cutoff),
                int(row["inspection_record_count"]),
            ]
        )

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()
