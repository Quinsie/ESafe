from datetime import UTC, date, datetime, timedelta
from io import BytesIO
from uuid import uuid4

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app.api.auth import require_session
from app.auth import AuthenticatedSession
from app.data_extract import EXPORT_HEADERS, build_extract_workbook
from app.main import app
from app.security import token_hash


def _session() -> AuthenticatedSession:
    return AuthenticatedSession(
        session_id_hash=token_hash("session"),
        user_id=uuid4(),
        username="user",
        display_name="사용자",
        csrf_token_hash=token_hash("csrf"),
        expires_at=datetime.now(UTC) + timedelta(hours=1),
    )


def _row(
    *,
    rank: int,
    last_inspection_date: date | None,
) -> dict[str, object]:
    return {
        "building_name": f"건물 {rank}",
        "road_address": f"광주광역시 도로 {rank}",
        "lot_address": f"광주광역시 지번 {rank}",
        "sido_name": "광주광역시",
        "sigungu_name": "북구",
        "building_id": f"building-{rank}",
        "selected_region_rank": rank,
        "regional_rank": rank + 10,
        "final_score": 0.99 - rank / 100,
        "top_percentile": float(rank),
        "main_use_name": "공동주택",
        "latest_inspection_date": last_inspection_date,
        "inspection_record_count": rank + 2,
    }


def test_data_extract_endpoints_require_authentication() -> None:
    with TestClient(app) as client:
        regions = client.get("/api/v1/data-extract/regions?level=SIDO")
        workbook = client.get(
            "/api/v1/data-extract/buildings.xlsx"
            "?level=SIDO&regionCode=29&topPercent=10"
        )

    assert regions.status_code == 401
    assert workbook.status_code == 401


def test_extract_workbook_has_fourteen_columns_and_inspection_windows() -> None:
    content = build_extract_workbook(
        level="SIGUNGU",
        region_name="광주광역시 북구",
        top_percent=10,
        rows=[
            _row(rank=1, last_inspection_date=date(2026, 5, 1)),
            _row(rank=2, last_inspection_date=date(2025, 10, 1)),
            _row(rank=3, last_inspection_date=None),
        ],
        as_of=date(2026, 7, 30),
    )

    workbook = load_workbook(BytesIO(content), data_only=True)
    assert workbook.sheetnames == ["추출 조건", "건축물 목록"]
    sheet = workbook["건축물 목록"]
    assert sheet.max_column == 14
    assert sheet.auto_filter.ref == "A1:N4"
    assert tuple(cell.value for cell in next(sheet.iter_rows(max_row=1))) == EXPORT_HEADERS
    assert "도로명주소" not in EXPORT_HEADERS
    assert "건물 식별번호" not in EXPORT_HEADERS
    values = list(sheet.iter_rows(min_row=2, values_only=True))
    assert values[0][2] == "광주광역시 지번 1"
    assert values[0][9] == "공동주택"
    assert values[0][11:13] == ("있음", "있음")
    assert values[1][11:13] == ("없음", "있음")
    assert values[2][11:13] == ("미등록", "미등록")


def test_data_extract_region_contract(monkeypatch) -> None:
    app.dependency_overrides[require_session] = _session
    captured: dict[str, object] = {}
    expected = {
        "level": "SIGUNGU",
        "levelName": "시·군·구",
        "items": [
            {
                "regionCode": "29170",
                "fullName": "광주광역시 북구",
                "eligibleCounts": {"1": 10, "5": 50, "10": 100},
            }
        ],
    }

    async def options(engine, level, parent_code, timeout_seconds):
        captured["level"] = level
        captured["parent_code"] = parent_code
        return expected

    monkeypatch.setattr("app.api.data_extract.extract_region_options", options)
    try:
        with TestClient(app) as client:
            response = client.get(
                "/api/v1/data-extract/regions?level=SIGUNGU&parentCode=29"
            )
            unsupported = client.get(
                "/api/v1/data-extract/regions?level=EUPMYEONDONG&parentCode=29170"
            )
        assert response.status_code == 200
        assert unsupported.status_code == 422
        assert response.json()["data"] == expected
        assert captured == {"level": "SIGUNGU", "parent_code": "29"}
    finally:
        app.dependency_overrides.clear()


def test_data_extract_download_accepts_query_string_percent(monkeypatch) -> None:
    app.dependency_overrides[require_session] = _session
    captured: dict[str, object] = {}

    async def rows(engine, level, region_code, top_percent, timeout_seconds):
        captured["level"] = level
        captured["region_code"] = region_code
        captured["top_percent"] = top_percent
        return "광주광역시", []

    monkeypatch.setattr("app.api.data_extract.extract_building_rows", rows)
    monkeypatch.setattr("app.api.data_extract.build_extract_workbook", lambda **kwargs: b"PKxlsx")
    try:
        with TestClient(app) as client:
            response = client.get(
                "/api/v1/data-extract/buildings.xlsx"
                "?level=SIDO&regionCode=29&topPercent=10"
            )
        assert response.status_code == 200
        assert response.content == b"PKxlsx"
        assert captured == {
            "level": "SIDO",
            "region_code": "29",
            "top_percent": 10,
        }
    finally:
        app.dependency_overrides.clear()
